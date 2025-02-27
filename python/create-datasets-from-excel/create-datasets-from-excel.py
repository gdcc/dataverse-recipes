"""
Create Dataverse datasets from an Excel spreadsheet.

This script reads data from a 'data.xlsx' Excel file and creates corresponding datasets
in a Dataverse collection. Each row in the spreadsheet becomes a separate dataset.

Usage:
    1. Place your data in 'data.xlsx' in the same directory as this script
    2. Set environment variables (optional):
        - SERVER_URL: URL of your Dataverse instance (default: http://localhost:8080)
        - API_TOKEN: Your Dataverse API token (default: 3e2b87bf-6652-47f1-b0bb-ddc0aa1a9106)
        - COLLECTION: Name of collection to create datasets in (default: collection1)
    3. Run the script: python create-datasets-from-excel.py

The Excel file should contain columns matching the metadata fields you want to populate
in your datasets. Special characters will be automatically cleaned during import.

Requirements:
    - easyDataverse
    - openpyxl
"""

import time
import os
import json

from easyDataverse import Dataverse

# we use openpyxl because it can pull hyperlinks out of Excel
from openpyxl import load_workbook

###### Environment variables ######

# SERVER_URL="https://demo.dataverse.org"
SERVER_URL = str(
    os.environ.get(
        "SERVER_URL",
        "http://localhost:8080",  # default if not set
    )
)
API_TOKEN = str(
    os.environ.get(
        "API_TOKEN",
        "3e2b87bf-6652-47f1-b0bb-ddc0aa1a9106",  # default if not set
    )
)

# The collection where the datasets will be created
COLLECTION = str(
    os.environ.get(
        "COLLECTION",
        "collection1",  # default if not set
    )
)


###### Functions ######


def clean(value):
    """Helper function to replace special characters with standard ones"""
    if value is None:
        return ""
    if isinstance(value, str):
        # Replace special characters with standard ones
        return value.replace("’", "'").replace("–", "-")
    return str(value)


###### Main script ######

wb = load_workbook("data.xlsx")
sheets = wb.sheetnames
ws = wb[sheets[1]]

# Connect to a Dataverse installation
dataverse = Dataverse(
    server_url=SERVER_URL,
    api_token=API_TOKEN,
)

# min_row=2 to skip the header row
count = 0
for key, *values in ws.iter_rows(min_row=2):
    # Strangely, datasets created with EasyDataverse have
    # Custom Terms rather than CC0, which is the default,
    # and it's not possible to change the license afterward
    # using EasyDataverse: https://github.com/gdcc/easyDataverse/issues/29
    # So, we create a basic dataset using requests and
    # then update the fields later with EasyDataverse.

    with open("initial-dataset.json") as f:
        # Use the native API to create the dataset
        response = dataverse.native_api.create_dataset(
            dataverse=COLLECTION,
            metadata=json.load(f),
        )

        # Catch errors early
        response.raise_for_status()

        # We will need this pid later to update the dataset
        pid = response.json()["data"]["persistentId"]

    # Get the fields from Excel
    title = clean(values[0].value)
    description = clean(values[1].value)
    agency_responsible = clean(values[2].value)
    access_link_name = clean(values[3].value)

    if values[3].hyperlink:
        access_link_url = values[3].hyperlink.display
    else:
        # In case there is no hyperlink, we set the
        # access link url to None
        access_link_url = None

    print(f"Title: {title}")
    print(f"Description: {description}")
    print(f"Agency Responsible: {agency_responsible}")
    print(f"Access Link Name: {access_link_name}")
    print(f"Access Link URL: {access_link_url}")

    # Load the dataset we created earlier
    dataset = dataverse.load_dataset(pid)

    # Update the metadata fields
    dataset.citation.title = title
    dataset.citation.subject = ["Other"]
    dataset.metadatablocks["citation"].ds_description[0].value = description
    dataset.metadatablocks["citation"].author[0].name = agency_responsible
    dataset.metadatablocks["citation"].author[0].affiliation = None

    ds_contact = dataset.metadatablocks["citation"].dataset_contact[0]
    ds_contact.name = "Harvard Dataverse Support"
    ds_contact.email = "support@dataverse.harvard.edu"

    if access_link_url:
        dataset.metadatablocks["citation"].origin_of_sources = f'<a href="{access_link_url}">{access_link_name}</a>'

    # Update the dataset, pushing the new fields to the server
    dataset.update()

    # Sleep a bit before creating the next dataset
    time.sleep(2)

    # Uncomment the following line to create only one dataset
    # break
